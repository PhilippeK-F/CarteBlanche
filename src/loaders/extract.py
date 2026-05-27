import pandas as pd
from openpyxl import load_workbook
import warnings
warnings.filterwarnings('ignore')


def extract_patientele():
    """Extrait et nettoie la patientèle médecins traitants."""
    df = pd.read_csv(
        'data/raw/patientele-medecintraitant-generalistes-annuelle.csv',
        sep=';', encoding='utf-8-sig'
    )
    df.columns = [c.strip().lstrip('\ufeff') for c in df.columns]
    df = df.rename(columns={
        'annee': 'annee',
        'region': 'code_region',
        'libelle_region': 'libelle_region',
        'departement': 'code_dep',
        'libelle_departement': 'libelle_dep',
        'nombre_patients_medecin_traitant': 'nombre_patients',
        'taux_evolution_annuel': 'taux_evolution',
    })

    # Garder uniquement les départements réels
    df = df[df['code_dep'] != 999]
    df = df[df['code_dep'].notna()]
    df['code_dep'] = df['code_dep'].astype(str).str.zfill(3)
    df['nombre_patients'] = pd.to_numeric(df['nombre_patients'], errors='coerce')
    df['annee'] = pd.to_numeric(df['annee'], errors='coerce').astype('Int64')
    df = df.dropna(subset=['annee', 'code_dep', 'nombre_patients'])

    df.to_csv('data/processed/patientele_clean.csv', index=False, encoding='utf-8-sig')
    print(f"Patientèle : {len(df)} lignes, {df['annee'].nunique()} années, {df['code_dep'].nunique()} départements")
    return df


def extract_apl():
    """Extrait et nettoie les données APL par commune."""
    wb = load_workbook('data/raw/Indicateur_APL.xlsx', read_only=True)

    frames = []
    for sheet_name in ['APL 2022', 'APL 2023']:
        year = int(sheet_name.split()[1])
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Trouver la ligne d'en-têtes
        header_idx = None
        for i, row in enumerate(rows):
            if row[0] is not None and str(row[0]).strip().lower() in ['codgeo', 'code_commune', 'code commune']:
                header_idx = i
                break
            if row[0] is not None and len(str(row[0])) <= 6 and str(row[0]).strip().isdigit():
                header_idx = i - 1
                break

        if header_idx is None:
            for i, row in enumerate(rows):
                vals = [v for v in row if v is not None]
                if len(vals) >= 3 and any('apl' in str(v).lower() for v in vals):
                    header_idx = i
                    break

        if header_idx is not None:
            headers = [str(v).strip() if v is not None else f'col_{j}' for j, v in enumerate(rows[header_idx])]
            data = [dict(zip(headers, row)) for row in rows[header_idx+1:] if any(v is not None for v in row)]
            df = pd.DataFrame(data)
            df['annee'] = year
            frames.append(df)
            print(f"APL {year}: {len(df)} communes — colonnes: {headers[:6]}")

    if not frames:
        print("Erreur: structure APL non reconnue")
        return pd.DataFrame()

    return pd.concat(frames, ignore_index=True)


def extract_dim_departement(df_patientele):
    """Extrait la table de dimension départements."""
    df = df_patientele[['code_dep', 'libelle_dep', 'code_region', 'libelle_region']].drop_duplicates()
    df = df[df['code_dep'] != '999']
    df.to_csv('data/processed/dim_departement.csv', index=False, encoding='utf-8-sig')
    print(f"Dim départements : {len(df)} lignes")
    return df


if __name__ == "__main__":
    print("Extraction patientèle...")
    df_pat = extract_patientele()

    print("\nExtraction APL...")
    df_apl = extract_apl()

    print("\nExtraction dim départements...")
    df_dep = extract_dim_departement(df_pat)

    print("\n✅ Extraction terminée")